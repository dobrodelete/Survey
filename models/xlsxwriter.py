import os
import json

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Color, PatternFill

from config import MAIN_REPORTS_FOLDER, XLSX_REPORTS_FOLDER, JSON_REPORTS_FOLDER


def writedata(d):
    r = 1
    while True:
        if os.path.exists(
                f"{MAIN_REPORTS_FOLDER}/{JSON_REPORTS_FOLDER}/{d['subdivision'].replace(' ', '_')}_{d['post'].replace(' ', '_')}{r}.json") is False:
            with open(
                    f"{MAIN_REPORTS_FOLDER}/{JSON_REPORTS_FOLDER}/{d['subdivision'].replace(' ', '_')}_{d['post'].replace(' ', '_')}{r}.json",
                    "w") as file:
                json.dump(d, file)
            break
        r += 1
    print(d)
    with open("temp.json", "w") as file:
        json.dump(d, file)
    insert_data = create_insert_data_list()
    print(os.getcwd())

    wb = load_workbook(filename="models/template.xlsx")
    for data in insert_data:
        ws = wb[data]
        for quest in d["answer_questions"]:
            for q, r in insert_data[data].items():
                if int(quest["question"]) == int(q):
                    print(f'{data} : {q} : {r} : {float(quest["reply"]["answer"])}')
                    ws[r].value = float(quest["reply"]["answer"])
                    if "comment" in quest["reply"]:
                        comment = Comment(quest["reply"]["comment"], "admin")
                        comment.width = 300
                        comment.height = 50
                        ws[r].comment = comment
                        yellow = PatternFill(start_color='FFFF00',
                                              end_color='FFFF00',
                                              fill_type='solid')
                        ws[r].fill = yellow
    flag = True
    t = 1
    while flag:
        try:
            if not os.path.exists(f"{MAIN_REPORTS_FOLDER}/{XLSX_REPORTS_FOLDER}/{d['iogv_id']}"):
                os.mkdir(f"{MAIN_REPORTS_FOLDER}/{XLSX_REPORTS_FOLDER}/{d['iogv_id']}")
            wb.save(f"{MAIN_REPORTS_FOLDER}/{XLSX_REPORTS_FOLDER}/{d['iogv_id']}/{d['subdivision'].replace(' ', '_')}_{d['post'].replace(' ', '_')}{t}.xlsx")
            flag = False
        except Exception as ex:
            print(ex)
        t += 1
        if t == 15:
            flag = False


def create_insert_data_list():
    data = dict()
    data["1. Культура и управление"] = {}
    data["2. Кадры"] = {}
    data["3. Процессы"] = {}
    data["4. Продукты"] = {}
    data["5. Данные"] = {}
    data["6. Инфраструктура и инструменты"] = {}
    i = 1
    for j in range(5, 18):
        data["1. Культура и управление"][f"{i}"] = f"E{j}"
        i += 1
    for j in range(5, 20):
        data["2. Кадры"][f"{i}"] = f"E{j}"
        i += 1
    for j in range(5, 14):
        data["3. Процессы"][f"{i}"] = f"E{j}"
        i += 1
    for j in range(5, 16):
        data["4. Продукты"][f"{i}"] = f"E{j}"
        i += 1
    for j in range(5, 17):
        data["5. Данные"][f"{i}"] = f"E{j}"
        i += 1
    for j in range(5, 11):
        data["6. Инфраструктура и инструменты"][f"{i}"] = f"E{j}"
        i += 1
    return data
